import json
import os
import re
import subprocess
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import streamlit as st

ROOT_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT_DIR / "python"))
from pc_excel_spec import REQUIRED_PC_COLUMNS

PC_XLSX = ROOT_DIR / "python" / "data" / "uk" / "pc_data.xlsx"
PC_JSON = ROOT_DIR / "web" / "public" / "uk" / "pc_data.json"

DEFAULT_HEADER_ROW = 4
DEFAULT_PARENT_TENANT_ID = 15


def env_positive_int(name: str) -> int:
    raw = (os.getenv(name) or "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return 0


def excel_col_index_to_letters(index: int) -> str:
    if index <= 0:
        return "A"
    letters = []
    n = index
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def normalize_header(h: str) -> str:
    text = str(h).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def read_header_preview(file_bytes: bytes, header_row: int) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=header_row, max_row=header_row):
            out = []
            for cell in row:
                letter = excel_col_index_to_letters(cell.column or 0)
                value = "" if cell.value is None else str(cell.value).strip()
                out.append((letter, value or "(empty)"))
            return out
        return []
    finally:
        wb.close()


def match_required_headers(preview: list[tuple[str, str]]) -> dict[str, str]:
    by_norm: dict[str, str] = {}
    for letter, name in preview:
        if name == "(empty)":
            continue
        by_norm.setdefault(normalize_header(name), letter)
    found: dict[str, str] = {}
    for col in REQUIRED_PC_COLUMNS:
        for alias in col["aliases"]:
            letter = by_norm.get(normalize_header(alias))
            if letter:
                found[col["key"]] = letter
                break
    return found


def backup_existing_xlsx() -> Path | None:
    if not PC_XLSX.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = PC_XLSX.with_name(f"pc_data.{stamp}.xlsx")
    suffix = 2
    while dest.exists():
        dest = PC_XLSX.with_name(f"pc_data.{stamp}-{suffix}.xlsx")
        suffix += 1
    PC_XLSX.rename(dest)
    return dest


def read_product_count() -> int | None:
    if not PC_JSON.exists():
        return None
    try:
        payload = json.loads(PC_JSON.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    meta = payload.get("meta") if isinstance(payload, dict) else None
    if not isinstance(meta, dict):
        return None
    count = meta.get("count")
    return count if isinstance(count, int) else None


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    def cell(value: str) -> str:
        return str(value).replace("|", "\\|").replace("\n", " ")

    head = "| " + " | ".join(cell(h) for h in headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    body = ["| " + " | ".join(cell(c) for c in row) + " |" for row in rows]
    return "\n".join([head, sep, *body])


st.set_page_config(page_title="PC Excel uploader", layout="centered")
st.title("PC Excel uploader")
st.write(
    "Drop the spreadsheet, then enter the **header row** (the row with column names). "
    "Every listed header must exist. Each row needs **DESCRIPTION** and **PRODUCT CODE**. "
    "HAZARDOUS = yes marks the product hazardous on sync. Pictures come from **IMAGE**. MOQ is always **6**. "
    "INNER CASE is still stored as case size. "
    "Warehouse parent tenant defaults to **15**."
)

uploaded = st.file_uploader("Drop a PC spreadsheet", type=["xlsx"])
header_row = st.number_input(
    "Header row (row with column names)",
    min_value=1,
    step=1,
    value=DEFAULT_HEADER_ROW,
    help="Excel row number that contains PRODUCT CODE, BARCODE, DESCRIPTION, and the other required headers.",
)

st.subheader("Required headers")
st.markdown(
    markdown_table(
        ["Excel header", "Goes to", "Meaning"],
        [
            [
                col["excel"],
                col["db"].replace("products.", ""),
                col.get("note") or "Required",
            ]
            for col in REQUIRED_PC_COLUMNS
        ],
    )
)

preview = None
found_headers: dict[str, str] = {}
if uploaded is not None:
    try:
        preview = read_header_preview(uploaded.getvalue(), int(header_row))
    except Exception as exc:
        st.error(f"Could not read headers: {exc}")
        preview = None
    if preview:
        found_headers = match_required_headers(preview)
        st.subheader("This file")
        st.markdown(
            markdown_table(
                ["Excel header", "Column", "Status"],
                [
                    [
                        col["excel"],
                        found_headers.get(col["key"]) or "—",
                        "Found" if col["key"] in found_headers else "Missing",
                    ]
                    for col in REQUIRED_PC_COLUMNS
                ],
            )
        )
        missing = [col["excel"] for col in REQUIRED_PC_COLUMNS if col["key"] not in found_headers]
        if missing:
            st.error("Missing required header(s): " + ", ".join(missing))

parent_tenant_id = st.number_input(
    "Parent tenant id",
    min_value=1,
    step=1,
    value=env_positive_int("PY_PRODUCTS_PARENT_TENANT_ID") or DEFAULT_PARENT_TENANT_ID,
    help="Warehouse HQ tenant. Default 15.",
)

run = st.button(
    "Save and run",
    type="primary",
    disabled=uploaded is None or st.session_state.get("running", False),
)
log_box = st.empty()

if run and uploaded is not None:
    missing = [col["excel"] for col in REQUIRED_PC_COLUMNS if col["key"] not in found_headers]
    if int(parent_tenant_id) < 1:
        st.error("Parent tenant id is required.")
    elif missing:
        st.error("Missing required header(s): " + ", ".join(missing))
    else:
        st.session_state["running"] = True
        PC_XLSX.parent.mkdir(parents=True, exist_ok=True)
        backup_path = backup_existing_xlsx()
        PC_XLSX.write_bytes(uploaded.getvalue())

        flags = f"--header-row {int(header_row)}"
        sync_flags = f"--parent-tenant-id {int(parent_tenant_id)}"
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["PY_PRODUCTS_PARENT_TENANT_ID"] = str(int(parent_tenant_id))
        lines: list[str] = []
        if backup_path:
            lines.append(f"Backed up previous sheet to {backup_path.name}\n")
        lines.append(f"Saved {PC_XLSX.name}\n")
        lines.append(
            f"Running: make -C python pc PC_EXPORT_FLAGS=\"{flags}\" "
            f"PC_SYNC_FLAGS=\"{sync_flags}\"\n\n"
        )
        log_box.code("".join(lines))

        proc = subprocess.Popen(
            [
                "make",
                "-C",
                str(ROOT_DIR / "python"),
                "pc",
                f"PC_EXPORT_FLAGS={flags}",
                f"PC_SYNC_FLAGS={sync_flags}",
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            env=env,
        )
        assert proc.stdout is not None
        for line in proc.stdout:
            lines.append(line)
            log_box.code("".join(lines[-400:]))
        code = proc.wait()
        st.session_state["running"] = False
        if code == 0:
            count = read_product_count()
            if count is not None:
                st.success(f"Done. Product count: {count}")
            else:
                st.success("Done.")
        else:
            st.error(f"Pipeline failed (exit {code}).")
