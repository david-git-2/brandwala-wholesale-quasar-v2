import os
import re
import json
import time
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import requests

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow


ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
ROOT_ENV_FILE = os.path.join(ROOT_DIR, ".env")
WEB_ENV_FILE = os.path.join(ROOT_DIR, "web", ".env")


def load_env_file(path: str):
    """Minimal .env loader without extra dependencies."""
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[7:].strip()
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if (
                len(value) >= 2
                and ((value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")))
            ):
                value = value[1:-1]
            if key:
                os.environ.setdefault(key, value)


# Shared env strategy:
# 1) Prefer web/.env as single source for web + python settings.
# 2) Fall back to root .env for backward compatibility.
load_env_file(WEB_ENV_FILE)
load_env_file(ROOT_ENV_FILE)

def env_path(env_key: str, default_abs_path: str) -> str:
    raw = str(os.getenv(env_key, "")).strip()
    if not raw:
        return default_abs_path
    return raw if os.path.isabs(raw) else os.path.join(ROOT_DIR, raw)


def env_positive_int(env_key: str, default: int) -> int:
    raw = str(os.getenv(env_key, "")).strip()
    if not raw:
        return default
    try:
        value = int(raw)
        return value if value > 0 else default
    except ValueError:
        return default


DEFAULT_XLSX = env_path("PY_UK_PC_XLSX_PATH", os.path.join(ROOT_DIR, "python", "data", "uk", "pc_data.xlsx"))
DEFAULT_OUT_JSON = env_path("PY_UK_PC_OUT_JSON_PATH", os.path.join(ROOT_DIR, "web", "public", "uk", "pc_data.json"))
DEFAULT_OUT_MANIFEST = env_path("PY_UK_PC_OUT_MANIFEST_PATH", os.path.join(ROOT_DIR, "web", "public", "uk", "pc_manifest.json"))
DEFAULT_OUT_IMAGES = env_path("PY_UK_PC_OUT_IMAGES_DIR", os.path.join(ROOT_DIR, "python", "images", "uk", "out_images"))
DEFAULT_SUPABASE_BUCKET = os.getenv("PY_SUPABASE_STORAGE_BUCKET", "product-images").strip() or "product-images"
DEFAULT_SUPABASE_PREFIX = os.getenv("PY_SUPABASE_STORAGE_PREFIX", "uk/pc").strip().strip("/")
DEFAULT_UPLOAD_WORKERS = env_positive_int("PY_UK_PC_UPLOAD_WORKERS", 20)

CREDS_DIR = env_path("PY_GOOGLE_CREDS_DIR", os.path.join(ROOT_DIR, "python", "credentials"))
OAUTH_CLIENT_JSON = env_path("PY_GOOGLE_OAUTH_CLIENT_JSON", os.path.join(CREDS_DIR, "oauth_client.json"))
TOKEN_JSON = env_path("PY_GOOGLE_TOKEN_JSON", os.path.join(CREDS_DIR, "token.json"))
DEFAULT_DRIVE_FOLDER_ID = os.getenv("PY_GOOGLE_DRIVE_FOLDER_ID", "").strip()

SCOPES = ["https://www.googleapis.com/auth/drive.file"]  # only files this app creates


def log(msg: str):
    print(msg, flush=True)


def safe_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"[^\w\-\.]+", "_", s)
    return s[:120] if len(s) > 120 else s


def build_image_identity(product_code_value, barcode_value, row_number: int) -> tuple[str, str]:
    """
    Build stable identifiers for image/product keys.
    - Search/upload key uses barcode first: <barcode>__<product_code>
    - product_id uses barcode first: <barcode>_<product_code>
    """
    pc_str = safe_filename(product_code_value if product_code_value is not None else "")
    bc_str = safe_filename(barcode_value if barcode_value is not None else "")

    if not bc_str:
        bc_str = f"NO_BARCODE_ROW_{row_number}"
    if not pc_str:
        pc_str = f"NO_CODE_ROW_{row_number}"

    image_key = f"{bc_str}__{pc_str}".strip("_")
    product_id = f"{bc_str}_{pc_str}".strip("_")
    return image_key, product_id


def prompt_int(label: str, default: int) -> int:
    """Prompt user for an integer. If empty input, returns default."""
    while True:
        raw = input(f"{label} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            val = int(raw)
            if val <= 0:
                log("❌ Please enter a positive integer.")
                continue
            return val
        except ValueError:
            log("❌ Invalid number. Try again.")


def normalize_header(h: str) -> str:
    # Normalize for resilient header matching across spaces/symbols/casing.
    text = str(h).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def excel_col_index_to_letters(index: int) -> str:
    if index <= 0:
        return "A"
    letters = []
    n = index
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def excel_col_letters_to_index(value: str) -> int:
    s = str(value or "").strip().upper()
    if not s or not re.fullmatch(r"[A-Z]+", s):
        return 0
    result = 0
    for ch in s:
        result = (result * 26) + (ord(ch) - ord("A") + 1)
    return result


def prompt_excel_column(label: str, default_index: int) -> int:
    default_letters = excel_col_index_to_letters(default_index)
    while True:
        raw = input(f"{label} [{default_letters}]: ").strip()
        if raw == "":
            return default_index

        # Backward-compatible: allow numeric input too.
        if raw.isdigit():
            val = int(raw)
            if val > 0:
                return val
            log("❌ Please enter a valid column (A, B, C... or positive number).")
            continue

        col_idx = excel_col_letters_to_index(raw)
        if col_idx > 0:
            return col_idx

        log("❌ Invalid column. Use Excel letters like A, B, C, AA ...")


def resolve_header_name(header_to_col: dict, headers: list, candidates: list) -> str:
    """Find first matching header name by normalized candidate keys."""
    for candidate in candidates:
        col = header_to_col.get(normalize_header(candidate))
        if col:
            return headers[col - 1]
    return ""


def to_text(value) -> str:
    return str(value if value is not None else "").strip()


def to_float_or_default(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if text in ("", "-", ".", "-."):
        return default
    try:
        return float(text)
    except ValueError:
        return default


def to_int_or_default(value, default: int = 0) -> int:
    return int(round(to_float_or_default(value, float(default))))


def format_optional_expire_date(value) -> str:
    """Normalize optional expire date field; empty string when missing."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text


def normalize_yes_no_value(value) -> str:
    """Normalize spreadsheet yes/no style values for filtering."""
    text = to_text(value).lower()
    return text.replace(" ", "")


def is_hazardous_yes(value) -> bool:
    normalized = normalize_yes_no_value(value)
    return normalized in {"yes", "y", "true", "1"}


def get_oauth_credentials() -> Credentials:
    """
    Loads token.json if present, otherwise performs OAuth login.
    Refreshes expired tokens and saves back to token.json.
    Returns a valid Credentials object.
    """
    os.makedirs(CREDS_DIR, exist_ok=True)

    creds = None
    if os.path.exists(TOKEN_JSON):
        log("🔐 Found token.json, using existing login...")
        creds = Credentials.from_authorized_user_file(TOKEN_JSON, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            log("🔄 Token expired, refreshing...")
            creds.refresh(Request())
        else:
            if not os.path.exists(OAUTH_CLIENT_JSON):
                raise FileNotFoundError(
                    f"Missing OAuth client file: {OAUTH_CLIENT_JSON}\n"
                    f"Download 'Desktop app' OAuth client JSON and save it there."
                )
            log("🌐 Opening browser for Google login/consent...")
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CLIENT_JSON, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_JSON, "w", encoding="utf-8") as token:
            token.write(creds.to_json())
        log(f"✅ Saved token: {TOKEN_JSON}")

    return creds


def build_drive_service(creds: Credentials):
    """Build a Drive API client."""
    return build("drive", "v3", credentials=creds)


def drive_file_key_from_name(name: str) -> str:
    """
    Returns lower-cased basename without extension to keep matching resilient.
    Example: 123__abc.jpg -> 123__abc
    """
    base = os.path.basename(str(name or ""))
    stem, _ = os.path.splitext(base)
    return stem.strip().lower()


def list_drive_files_index(service, folder_id: str) -> dict:
    """
    Build a cache of existing files in the target folder:
      - by_name: exact filename -> file_id
      - by_key:  stem-lower (no extension) -> file_id
    If duplicates exist, latest modified file wins.
    """
    by_name = {}
    by_key = {}
    page_token = None
    query_parts = ["trashed = false"]
    if folder_id:
        query_parts.append(f"'{folder_id}' in parents")
    q = " and ".join(query_parts)

    while True:
        resp = (
            service.files()
            .list(
                q=q,
                fields="nextPageToken, files(id,name,modifiedTime)",
                pageSize=1000,
                pageToken=page_token,
            )
            .execute()
        )
        files = resp.get("files", [])
        # Stable ordering: older first, newer last, so newer wins on assignment.
        files.sort(key=lambda f: str(f.get("modifiedTime", "")))
        for f in files:
            fid = str(f.get("id", "")).strip()
            name = str(f.get("name", "")).strip()
            if not fid or not name:
                continue
            by_name[name] = fid
            by_key[drive_file_key_from_name(name)] = fid
        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return {"by_name": by_name, "by_key": by_key}


def upload_to_drive(service, local_path: str, folder_id: str, existing_file_id: str | None = None) -> tuple[str, str, str]:
    """
    Uploads or updates a file in Drive.
    IMPORTANT: Does NOT set per-file public permissions.
    You said your target folder is already public, so we skip permissions for speed.
    Returns: (direct_link, file_id, action[created|updated])
    """
    filename = os.path.basename(local_path)
    media = MediaFileUpload(local_path, resumable=True)
    if existing_file_id:
        updated = (
            service.files()
            .update(fileId=existing_file_id, media_body=media, fields="id")
            .execute()
        )
        file_id = updated["id"]
        action = "updated"
    else:
        metadata = {"name": filename}
        if folder_id:
            metadata["parents"] = [folder_id]
        created = service.files().create(body=metadata, media_body=media, fields="id").execute()
        file_id = created["id"]
        action = "created"

    # Direct link works well for <img src=""> if the file is accessible via folder sharing.
    direct_link = f"https://drive.google.com/uc?id={file_id}"
    return direct_link, file_id, action


def format_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h}h {m}m {s}s"
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def supabase_public_url(base_url: str, bucket: str, object_path: str) -> str:
    base = base_url.rstrip("/")
    return f"{base}/storage/v1/object/public/{bucket}/{object_path}"


def upload_to_supabase_storage(
    supabase_url: str,
    api_key: str,
    bucket: str,
    object_path: str,
    local_path: str,
) -> tuple[str, str]:
    """
    Upload/replace object in Supabase Storage.
    Returns (public_url, action) where action is 'created_or_updated'.
    """
    base = supabase_url.rstrip("/")
    url = f"{base}/storage/v1/object/{bucket}/{object_path}"
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "x-upsert": "true",
        "Content-Type": "application/octet-stream",
    }
    with open(local_path, "rb") as f:
        resp = requests.post(url, headers=headers, data=f, timeout=120)
    if not resp.ok:
        raise RuntimeError(f"supabase upload failed ({resp.status_code}): {resp.text}")
    return supabase_public_url(base, bucket, object_path), "created_or_updated"


def main():
    t0 = time.perf_counter()

    # ---- CONFIG ----
    EXCEL_PATH = DEFAULT_XLSX
    OUT_JSON_PATH = DEFAULT_OUT_JSON
    OUT_IMAGES_DIR = DEFAULT_OUT_IMAGES

    SUPABASE_URL = str(os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL") or "").strip()
    SUPABASE_ADMIN_KEY = str(
        os.getenv("SUPABASE_SECRET_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY") or ""
    ).strip()
    SUPABASE_BUCKET = DEFAULT_SUPABASE_BUCKET
    SUPABASE_PREFIX = DEFAULT_SUPABASE_PREFIX
    if not SUPABASE_URL:
        raise ValueError(
            "Missing SUPABASE_URL or VITE_SUPABASE_URL in web/.env."
        )
    if not SUPABASE_ADMIN_KEY:
        raise ValueError(
            "Missing SUPABASE_SECRET_KEY (preferred) or SUPABASE_SERVICE_ROLE_KEY (legacy) in web/.env."
        )

    DEFAULT_HEADER_ROW = 4
    DEFAULT_IMAGE_COLUMN_INDEX = 14  # 1-based column index (14 = N)

    SHEET_NAME = None

    # Parallel upload settings
    MAX_WORKERS = DEFAULT_UPLOAD_WORKERS
    # ----------------

    log(
        "\n"
        "🚀 PC export starting...\n"
        "\n"
        "📌 Excel format requirements:\n"
        "  Required logical fields in the HEADER ROW:\n"
        "    - product_code (e.g. PRODUCT CODE)\n"
        "    - barcode (e.g. BARCODE)\n"
        "    - case_size (e.g. CASE SIZE / INNER CASE)\n"
        "    - name (e.g. NAME / DESCRIPTION)\n"
        "    - price (e.g. PRICE / PIECE PRICE £)\n"
        "    - image   (images are embedded in the sheet; you will enter the image column letter)\n"
        "  Optional columns:\n"
        "    - country_of_origin\n"
        "    - brand\n"
        "    - expire_date\n"
        "    - hazardous (rows with YES are skipped)\n"
        "\n"
        "🆔 Product ID:\n"
        "  - product_id = barcode + '_' + product_code\n"
        "\n"
        "⚡ Speed mode enabled:\n"
        "  - Supabase Storage upsert enabled (same filename replaces old image).\n"
        f"  - Parallel uploads enabled (workers={MAX_WORKERS}).\n"
        "\n"
        "👉 You will be asked for:\n"
        "  - Header row number (where the column names are)\n"
        "  - Image column letter (A, B, C, ...)\n"
    )

    HEADER_ROW = prompt_int("Enter header row number", DEFAULT_HEADER_ROW)
    IMAGE_COLUMN_INDEX = prompt_excel_column("Enter image column letter (A, B, C, ...)", DEFAULT_IMAGE_COLUMN_INDEX)

    log(f"\n📄 Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    os.makedirs(OUT_IMAGES_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(OUT_JSON_PATH), exist_ok=True)
    log(f"📁 Output images folder: {OUT_IMAGES_DIR}")
    log(f"🧾 Output JSON: {OUT_JSON_PATH}\n")

    log("📥 Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sh = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]
    log(f"✅ Using sheet: {sh.title}")

    max_col = sh.max_column
    max_row = sh.max_row
    log(f"📐 Sheet size: rows={max_row}, cols={max_col}")

    # Read headers
    log(f"🏷️ Reading headers from row {HEADER_ROW}...")
    headers = []
    for c in range(1, max_col + 1):
        v = sh.cell(HEADER_ROW, c).value
        headers.append(str(v).strip() if v is not None else f"col_{c}")

    # Build header lookup (normalized -> column index)
    header_to_col = {}
    for idx, h in enumerate(headers, start=1):
        nh = normalize_header(h)
        header_to_col.setdefault(nh, idx)

    # Resolve logical fields from flexible header aliases.
    field_candidates = {
        "product_code": ["product_code", "product code", "code"],
        "barcode": ["barcode", "bar code", "ean"],
        "case_size": ["case_size", "case size", "inner_case", "inner case", "sales_unit", "sales unit"],
        "name": ["name", "description", "product_name", "product name"],
        "price": ["price", "piece_price", "piece price", "piece_price_gbp", "unit_price", "unit price"],
        "country_of_origin": ["country_of_origin", "country of origin", "country"],
        "brand": ["brand"],
        "expire_date": ["expire_date", "expiry_date", "expiry date", "expiration_date", "exp_date"],
        "category": ["category"],
        "available_units": ["available_units", "available units"],
        "inner_case": ["inner_case", "inner case"],
        "outer_case": ["outer_case", "outer case"],
        "outer_per_plt": ["outer_per_plt", "outer per plt", "outer per pallet"],
        "tariff_code": ["tariff_code", "tariff code"],
        "languages": ["languages", "language"],
        "batch_code_manufacture_date": [
            "batch_code_manufacture_date",
            "batch code / manufacture date",
            "batch code manufacture date",
            "batch_code",
            "manufacture_date",
        ],
        "hazardous": ["hazardous"],
        "sales_unit": ["sales_unit", "sales unit"],
        "image": ["image", "image url", "photo"],
    }
    resolved_headers = {
        field: resolve_header_name(header_to_col, headers, candidates)
        for field, candidates in field_candidates.items()
    }

    required = ["product_code", "barcode", "case_size", "name", "price"]
    missing_required = [field for field in required if not resolved_headers.get(field)]
    if missing_required:
        raise RuntimeError(
            "❌ Missing required column(s) in header row "
            f"{HEADER_ROW}: {', '.join(missing_required)}\n"
            "Make sure your Excel header row contains matching columns (case-insensitive).\n"
            "Examples: PRODUCT CODE, BARCODE, INNER CASE, DESCRIPTION, PIECE PRICE £\n"
        )

    missing_optional = [
        field
        for field in ["country_of_origin", "brand", "expire_date", "category", "hazardous"]
        if not resolved_headers.get(field)
    ]
    if missing_optional:
        log(f"ℹ️ Optional column(s) missing (OK): {', '.join(missing_optional)}")

    product_code_header_name = resolved_headers["product_code"]
    barcode_header_name = resolved_headers["barcode"]
    case_size_header_name = resolved_headers["case_size"]
    name_header_name = resolved_headers["name"]
    price_header_name = resolved_headers["price"]
    country_of_origin_header_name = resolved_headers.get("country_of_origin", "")
    brand_header_name = resolved_headers.get("brand", "")
    expire_date_header_name = resolved_headers.get("expire_date", "")
    category_header_name = resolved_headers.get("category", "")
    hazardous_header_name = resolved_headers.get("hazardous", "")

    product_code_col = header_to_col[normalize_header(product_code_header_name)]
    barcode_col = header_to_col[normalize_header(barcode_header_name)]
    log(f"✅ Found product_code column at index: {product_code_col}")
    log(f"✅ Found barcode column at index: {barcode_col}")
    log(
        f"✅ Found case_size column: {case_size_header_name} | "
        f"name column: {name_header_name} | price column: {price_header_name}"
    )
    log(
        f"🖼️ Using image column: {excel_col_index_to_letters(IMAGE_COLUMN_INDEX)} "
        f"(index={IMAGE_COLUMN_INDEX})\n"
    )

    # Read rows
    log("📦 Reading product rows...")
    start_data_row = HEADER_ROW + 1
    products_by_row = {}

    for r in range(start_data_row, max_row + 1):
        row_vals = [sh.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        obj = {}
        for c, h in enumerate(headers, start=1):
            obj[h] = row_vals[c - 1]
        obj["_rowNumber"] = r
        products_by_row[r] = obj

    log(f"✅ Products loaded: {len(products_by_row)}")

    # Filter hazardous rows first so image upload only processes needed products.
    eligible_rows = set(products_by_row.keys())
    hazardous_removed_rows = 0
    if hazardous_header_name:
        eligible_rows = set()
        for row, obj in products_by_row.items():
            if is_hazardous_yes(obj.get(hazardous_header_name, "")):
                hazardous_removed_rows += 1
                continue
            eligible_rows.add(row)
        log(
            f"🧪 Hazardous filter: removed={hazardous_removed_rows}, "
            f"kept={len(eligible_rows)}"
        )

    # Extract images
    log(
        "🖼️ Finding embedded images anchored to column "
        f"{excel_col_index_to_letters(IMAGE_COLUMN_INDEX)} (index={IMAGE_COLUMN_INDEX})..."
    )
    images = getattr(sh, "_images", [])
    log(f"🖼️ Total images detected in sheet: {len(images)}")

    img_by_row = {}
    for img in images:
        anchor = img.anchor._from  # 0-based
        row = anchor.row + 1
        col = anchor.col + 1

        if col != IMAGE_COLUMN_INDEX:
            continue
        if row not in products_by_row:
            continue
        if row not in eligible_rows:
            continue

        img_bytes = img._data()
        ext = (getattr(img, "format", None) or "jpg").lower()
        if ext == "jpeg":
            ext = "jpg"

        pc_val = products_by_row[row].get(product_code_header_name, "")
        bc_val = products_by_row[row].get(barcode_header_name, "")
        img_key, _ = build_image_identity(pc_val, bc_val, row)
        img_by_row[row] = (img_key, ext, img_bytes)

    log(f"✅ Images matched to product rows: {len(img_by_row)}")

    # Save images locally (deduped by stable image key).
    log("💾 Saving images locally...")
    local_path_by_key = {}
    image_key_by_row = {}

    for i, (row, (img_key, ext, img_bytes)) in enumerate(img_by_row.items(), start=1):
        # One stable filename per business key. Repeated keys overwrite same local file.
        filename = f"{img_key}.{ext}"
        local_path = os.path.join(OUT_IMAGES_DIR, filename)

        with open(local_path, "wb") as f:
            f.write(img_bytes)

        local_path_by_key[img_key] = local_path
        image_key_by_row[row] = img_key
        if i % 25 == 0 or i == len(img_by_row):
            log(f"   ...saved {i}/{len(img_by_row)}")

    log(
        f"✅ Local images saved: rows={len(image_key_by_row)}, "
        f"unique_keys={len(local_path_by_key)}\n"
    )

    log("☁️ Preparing Supabase Storage upload config...")
    log(
        f"✅ Supabase ready: bucket={SUPABASE_BUCKET}"
        + (f", prefix={SUPABASE_PREFIX}" if SUPABASE_PREFIX else "")
    )

    def upload_one(image_key: str, path: str):
        filename = os.path.basename(path)
        object_path = f"{SUPABASE_PREFIX}/{filename}" if SUPABASE_PREFIX else filename

        last_err = None
        for attempt in range(1, 4):
            try:
                url, action = upload_to_supabase_storage(
                    SUPABASE_URL,
                    SUPABASE_ADMIN_KEY,
                    SUPABASE_BUCKET,
                    object_path,
                    path,
                )
                return image_key, url, action
            except Exception as e:
                last_err = e
                time.sleep(min(2.0, 0.4 * attempt))
        raise RuntimeError(f"upload failed after retries: {last_err}")

    # Parallel upload
    items = list(local_path_by_key.items())
    total = len(items)
    storage_url_by_key = {}
    updated_count = 0

    log(
        f"⬆️ Upserting {total} unique image(s) to Supabase Storage "
        f"(parallel workers={MAX_WORKERS})..."
    )
    up_start = time.perf_counter()

    done = 0
    failed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = [ex.submit(upload_one, image_key, path) for image_key, path in items]

        for fut in as_completed(futures):
            try:
                image_key, url, action = fut.result()
                storage_url_by_key[image_key] = url
                if action == "created_or_updated":
                    updated_count += 1
            except Exception as e:
                failed += 1
                log(f"❌ Upload failed: {e}")
            finally:
                done += 1

                if done % 25 == 0 or done == total:
                    elapsed = time.perf_counter() - up_start
                    rate = (done / elapsed) if elapsed > 0 else 0.0
                    remaining = total - done
                    eta = (remaining / rate) if rate > 0 else 0
                    log(
                        f"   ...uploaded {done}/{total} "
                        f"(fail={failed}) | avg {rate:.2f} files/sec | ETA {format_duration(eta)}"
                    )

    uploaded_count = len(storage_url_by_key)
    log(
        f"✅ Upload step done. URLs ready: {uploaded_count} "
        f"(upserted={updated_count}, failed={failed})\n"
    )

    # Build JSON (field names come from Excel header row)
    log("🧾 Building JSON payload...")
    products = []

    for row, obj in products_by_row.items():
        if row not in eligible_rows:
            continue

        out = {}

        pc_val = obj.get(product_code_header_name, "")
        bc_val = obj.get(barcode_header_name, "")

        img_key, product_id = build_image_identity(pc_val, bc_val, row)

        # Canonical keys expected by the web app (keep stable names).
        out["product_code"] = to_text(obj.get(product_code_header_name, ""))
        out["barcode"] = to_text(obj.get(barcode_header_name, ""))
        out["case_size"] = max(1, to_int_or_default(obj.get(case_size_header_name, 1), 1))
        out["name"] = to_text(obj.get(name_header_name, ""))
        out["price"] = to_float_or_default(obj.get(price_header_name, 0), 0)
        out["country_of_origin"] = to_text(obj.get(country_of_origin_header_name, "")) if country_of_origin_header_name else ""
        out["brand"] = to_text(obj.get(brand_header_name, "")) if brand_header_name else ""
        out["category"] = to_text(obj.get(category_header_name, "")) if category_header_name else ""

        # Additional normalized keys from the provided UK sheet.
        out["available_units"] = to_text(obj.get(resolved_headers.get("available_units", ""), "")) if resolved_headers.get("available_units") else ""
        out["inner_case"] = to_text(obj.get(resolved_headers.get("inner_case", ""), "")) if resolved_headers.get("inner_case") else ""
        out["outer_case"] = to_text(obj.get(resolved_headers.get("outer_case", ""), "")) if resolved_headers.get("outer_case") else ""
        out["outer_per_plt"] = to_text(obj.get(resolved_headers.get("outer_per_plt", ""), "")) if resolved_headers.get("outer_per_plt") else ""
        out["tariff_code"] = to_text(obj.get(resolved_headers.get("tariff_code", ""), "")) if resolved_headers.get("tariff_code") else ""
        out["languages"] = to_text(obj.get(resolved_headers.get("languages", ""), "")) if resolved_headers.get("languages") else ""
        out["batch_code_manufacture_date"] = (
            to_text(obj.get(resolved_headers.get("batch_code_manufacture_date", ""), ""))
            if resolved_headers.get("batch_code_manufacture_date")
            else ""
        )
        out["sales_unit"] = to_text(obj.get(resolved_headers.get("sales_unit", ""), "")) if resolved_headers.get("sales_unit") else ""
        if resolved_headers.get("image"):
            out["image"] = obj.get(resolved_headers["image"])

        # ✅ stable product id
        out["product_id"] = product_id

        # ✅ imageUrl (Supabase Storage public URL)
        out["imageUrl"] = storage_url_by_key.get(img_key)
        out["expire_date"] = format_optional_expire_date(
            obj.get(expire_date_header_name, "") if expire_date_header_name else ""
        )

        products.append(out)

    payload = {
        "meta": {
            "generatedAt": datetime.utcnow().isoformat() + "Z",
            "sourceFile": os.path.basename(EXCEL_PATH),
            "sheet": sh.title,
            "count": len(products),
            "imagesExtracted": len(image_key_by_row),
            "imagesUploaded": uploaded_count,
            "imagesUpserted": updated_count,
            "headerRow": HEADER_ROW,
            "imageColumnIndex": IMAGE_COLUMN_INDEX,
            "imageColumnLetter": excel_col_index_to_letters(IMAGE_COLUMN_INDEX),
            "parallelWorkers": MAX_WORKERS,
            "storageProvider": "supabase",
            "storageBucket": SUPABASE_BUCKET,
            "storagePrefix": SUPABASE_PREFIX,
            "note": "Supabase Storage upsert uses barcode+product_code key.",
            "productIdRule": "product_id = barcode + '_' + product_code",
            "hazardousRowsRemoved": hazardous_removed_rows,
        },
        "products": products,
    }

    with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # ✅ Write manifest for the web app's versioning/caching system
    manifest = {
        "version": datetime.utcnow().strftime("%Y%m%d%H%M%S"),
        "updated_at": payload["meta"]["generatedAt"],
        "count": len(products),
    }
    with open(DEFAULT_OUT_MANIFEST, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    t_total = time.perf_counter() - t0

    log("\n✅ Done")
    log(f"- Products: {len(products)}")
    log(
        f"- Images extracted: rows={len(image_key_by_row)}, "
        f"unique_keys={len(local_path_by_key)} -> {OUT_IMAGES_DIR}"
    )
    log(
        f"- Images upserted (with URL): {uploaded_count} "
        f"(upserted={updated_count})"
    )
    log(f"- Upload failures: {failed}")
    log(f"- JSON written: {OUT_JSON_PATH}")
    log(f"⏱️ Total time: {format_duration(t_total)}\n")


if __name__ == "__main__":
    main()
